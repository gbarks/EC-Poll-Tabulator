#!/usr/bin/env python3

# ==========================================================
#  ElloCoaster ballot generator
#  Author: Grant Barker
#
#  Requires Python 3
# ==========================================================

from __future__ import print_function

import sys

if sys.version_info[0] < 3:
    print("Program requires Python 3; running {0}.{1}".format(sys.version_info[0], sys.version_info[1]))
    sys.exit()

import re
import lxml
import argparse
import datetime
from bs4 import BeautifulSoup
from urllib.request import urlopen
from operator import itemgetter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# command line arguments
parser = argparse.ArgumentParser(description='Pull coaster info from RCDB list into .csv/.xlsx ballot')

def valid_year(s):
    try:
        return datetime.datetime.strptime(s, "%Y").strftime("%Y")
    except ValueError:
        msg = "Not a valid year: '{0}'.".format(s)
        raise argparse.ArgumentTypeError(msg)

parser.add_argument("-i", "--rcdblink", action="append", required=True,
                    help="RCDB input url (required) - can use multiple -i args")
parser.add_argument("-o", "--outballot", default="rcdb_ballot",
                    help="specify name of output [ballot].csv/.xlsx file")
parser.add_argument("-O", "--outdetails", default="rcdb_ballot_details",
                    help="specify name of output [details].csv/.xlsx file")
parser.add_argument("-c", "--combineTracks", action="store_true",
                    help="don't make separate coaster entries for multi-tracks")
parser.add_argument("-u", "--skipunknown", action="store_true",
                    help="skip all coasters named 'unknown'")
parser.add_argument("-d", "--skipnodate", action="store_true",
                    help="skip all coasters with nonspecific open/close date")
parser.add_argument("-k", "--skipkiddie", action="store_true",
                    help="skip all kiddie coasters")
parser.add_argument("-p", "--skippowered", action="store_true",
                    help="skip all powered coasters")
parser.add_argument("-a", "--skipalpine", action="store_true",
                    help="skip all alpine coasters")
parser.add_argument("-w", "--skipwackyworm", action="store_true",
                    help="skip all 'Big Apple/Wacky Worm' layouts")
parser.add_argument("-b", "--skipbutterfly", action="store_true",
                    help="skip all Butterfly (Kiddie/Family 'U Shuttle') layouts")
parser.add_argument("-t", "--skipnophotos", action="store_true",
                    help="skip all RCDB entries with no photos")
parser.add_argument("-y", "--skipwrongyear", action="store_true",
                    help="skip all coasters that did not operate in given year")
parser.add_argument("-Y", "--setyear", default=str(datetime.datetime.now().year),
                    help="set year for -y; defaults to current year", type=valid_year)
parser.add_argument("-v", "--verbose", action="count", default=0,
                    help="print data as it's processed; duplicate for more detail")

args = parser.parse_args()

# format output filenames
if args.outballot != "rcdb_ballot":
    if args.outdetails == "rcdb_ballot_details":
        args.outdetails = args.outballot + "_details"

def main():
    coasters = []
    rcdblink = args.rcdblink

    # counters for while loop
    i = 0
    j = len(rcdblink)

    while i < j:
        if args.verbose > 0:
            print("<<< Now checking " + rcdblink[i] + " >>>")

        # handle pages that are lists of coasters
        if is_list_page(rcdblink[i]):
            response = urlopen(rcdblink[i])
            html = response.read()
            soup = BeautifulSoup(html, 'lxml')

            table = soup.find('tbody')

            # iterate over all coasters listed on the page
            for tr in table.find_all('tr'):
                td = tr.find_all('td')[1]
                url = "https://rcdb.com" + td.find('a').get('href')

                c = parse_rcdb_page(url)
                if c is not None:
                    if isinstance(c, list):
                        for item in c:
                            coasters.append(item)
                    else:
                        coasters.append(c)

            # check if there's another list page to scrape in the footer
            rfoot = soup.find('div', attrs={'id':'rfoot'})
            if rfoot is not None:
                for a in rfoot.find_all('a'):
                    if a.text == ">>":
                        rcdblink.insert(i+1, "https://rcdb.com" + a.get('href'))

        # handle pages that are (presumably) individual coasters
        else:
            c = parse_rcdb_page(rcdblink[i])
            if c is not None:
                if isinstance(c, list):
                    for item in c:
                        coasters.append(item)
                else:
                    coasters.append(c)

        # increment while loop counter
        i += 1
        j = len(rcdblink)

    # sort coasters by country, park, and coaster name
    coasters = sorted(coasters, key=itemgetter('country', 'park', 'name'))

    # open .csv files
    csvballot = open(args.outballot + ".csv", "w")
    csvdetails = open(args.outdetails + ".csv", "w")
    csvballot.write("Rank,Ridden?,Name,Local Name,Park,Country,Location,ID\n")
    csvdetails.write("Rank,Ridden?,Name,Local Name,Park,Country,Full City Name," +
                    "ID,Full Location,State,City,Status,Opening Date,Closing Date," + 
                    "Type,Scale,Make,Model,Sub-Model,# of Tracks,RCDB URL\n")

    # create Excel workbooks for .xlsx files
    xlballot = Workbook()
    xlballot.active.title = args.outballot
    xldetails = Workbook()
    xldetails.active.title = args.outdetails
    xlballot.active.append(["Rank","Ridden?","Name","Local Name","Park","Country","Location","ID"])
    xldetails.active.append(["Rank","Ridden?","Name","Local Name","Park","Country","Full City Name",
                            "ID","Full Location","State","City","Status","Opening Date","Closing Date",
                            "Type","Scale","Make","Model","Sub-Model","# of Tracks","RCDB URL"])

    for c in coasters:

        # compose row for minimal .csv ballot
        csvline = "0,No,"
        csvline = none_to_blank(csvline, c, "name")
        csvline = none_to_blank(csvline, c, "altname")
        csvline = none_to_blank(csvline, c, "park")
        csvline = none_to_blank(csvline, c, "country")
        csvline = none_to_blank(csvline, c, "fullcity")
        csvline = csvline + c["id"] + "\n"
        csvballot.write(csvline)

        # compose row for detailed .csv ballot
        csvline = csvline[:-1] + ","
        csvline = none_to_blank(csvline, c, "location")
        csvline = none_to_blank(csvline, c, "state")
        csvline = none_to_blank(csvline, c, "city")
        csvline = none_to_blank(csvline, c, "status")
        csvline = none_to_blank(csvline, c, "opendate")
        csvline = none_to_blank(csvline, c, "closedate")
        csvline = none_to_blank(csvline, c, "type")
        csvline = none_to_blank(csvline, c, "scale")
        csvline = none_to_blank(csvline, c, "make")
        csvline = none_to_blank(csvline, c, "model")
        csvline = none_to_blank(csvline, c, "submodel")
        csvline = none_to_blank(csvline, c, "tracks")
        csvline = csvline + c["url"] + "\n"
        csvdetails.write(csvline)

        # compose row for minimal .xlsx ballot
        xlrow = ["0"]
        xlrow.append("No")
        xlrow.append(for_xl_output(c, "name"))
        xlrow.append(for_xl_output(c, "altname"))
        xlrow.append(for_xl_output(c, "park"))
        xlrow.append(for_xl_output(c, "country"))
        xlrow.append(for_xl_output(c, "fullcity"))
        xlrow.append('=HYPERLINK("{0}", "{1}")'.format(c["url"], c["id"]))
        xlballot.active.append(xlrow)

        # compose row for detailed .xlsx ballot
        xlrow.append(for_xl_output(c, "location"))
        xlrow.append(for_xl_output(c, "state"))
        xlrow.append(for_xl_output(c, "city"))
        xlrow.append(for_xl_output(c, "status"))
        xlrow.append(for_xl_output(c, "opendate"))
        xlrow.append(for_xl_output(c, "closedate"))
        xlrow.append(for_xl_output(c, "type"))
        xlrow.append(for_xl_output(c, "scale"))
        xlrow.append(for_xl_output(c, "make"))
        xlrow.append(for_xl_output(c, "model"))
        xlrow.append(for_xl_output(c, "submodel"))
        xlrow.append(for_xl_output(c, "tracks"))
        xlrow.append('=HYPERLINK("{0}", "{1}")'.format(c["url"], c["url"]))
        xldetails.active.append(xlrow)

    for i in range(len(coasters)):
        xlballot.active.cell(row=i+2, column=8).style = "Hyperlink"
        xldetails.active.cell(row=i+2, column=8).style = "Hyperlink"
        xldetails.active.cell(row=i+2, column=21).style = "Hyperlink"

    # close .csv files
    csvballot.close()
    csvdetails.close()

    # write .xlsx files
    xlballot.active.freeze_panes = xlballot.active['A2']
    xldetails.active.freeze_panes = xldetails.active['A2']
    xlballot.save(args.outballot + ".xlsx")
    xldetails.save(args.outdetails + ".xlsx")

def is_list_page(url):
    substring = url.split("rcdb.com/", 1)[1]
    if substring[:5] == "r.htm":
        return True
    else:
        return False

def none_to_blank(csvline, c, key):
    if key not in c or c[key] is None:
        return csvline + ","
    else:
        return csvline + c[key] + ","

def for_xl_output(c, key):
    if key not in c or c[key] is None:
        return ""
    else:
        string = c[key]
        if string[:1] == "\"" and string[-1:] == "\"":
            string = string[1:-1]
        return string

def parse_rcdb_page(url):
    c = {}
    c["url"] = url

    cresponse = urlopen(url)
    chtml = cresponse.read()
    csoup = BeautifulSoup(chtml, 'lxml')

    # get name, alt name (for native language users), park, and location
    feature = csoup.find('div', attrs={'id':'feature'})
    title = feature.find('div', attrs={'class':'scroll'})
    name = title.find('h1').text
    if " / " in name:
        altname = name.split(" / ", 1)[1]
        name = name.split(" / ", 1)[0]
        c["altname"] = "\"" + altname + "\""
    c["name"] = "\"" + name + "\""
    park = title.find_all('a')[0].text
    c["park"] = "\"" + park + "\""
    location = title.text[title.text.find("(")+1:title.text.find(")")]
    c["location"] = "\"" + location + "\""

    # get individual city name from the location string
    c["country"] = None
    c["fullcity"] = None
    c["state"] = None
    c["city"] = None
    if "," in location:
        country = location[location.rindex(",")+1:].strip()
        c["country"] = "\"" + country + "\""
        fullcity = location[:location.rindex(",")].strip()
        c["fullcity"] = "\"" + fullcity + "\""
        if "," in fullcity:
            state = fullcity[fullcity.rindex(",")+1:].strip()
            c["state"] = "\"" + state + "\""
            city = fullcity[:fullcity.rindex(",")].strip()
            c["city"] = "\"" + city + "\""

    # skip coasters with no header photo (-t arg)
    if args.skipnophotos and not csoup.find('a', attrs={'id':'opfAnchor'}):
        if args.verbose > 0:
            print("--Skipping " + name + " at " + park + " (No Photos)")
        return None

    # skip coasters named "Unknown" (-u arg)
    if args.skipunknown is True and name == "unknown":
        if args.verbose > 0:
            print("--Skipping \"unknown\" at " + park + " - " + location)
        return None

    # get opening date and closing date (and extract coaster type from date string)
    dates = feature.find_all('time')
    datestr = feature.text[feature.text.find(")")+1:]
    if "Mountain Coaster" in datestr:
        if args.skipalpine:
            if args.verbose > 0:
                print("--Skipping " + name + " at " + park + " (Alpine Coaster)")
            return None
        datestr = datestr[:datestr.find("Mountain Coaster")]
        c["type"] = "Mountain Coaster"
    elif "Powered Coaster" in datestr:
        if args.skippowered:
            if args.verbose > 0:
                print("--Skipping " + name + " at " + park + " (Powered Coaster)")
            return None
        datestr = datestr[:datestr.find("Powered Coaster")]
        c["type"] = "Powered Coaster"
    else:
        datestr = datestr[:datestr.find("Roller Coaster")]
        c["type"] = "Roller Coaster"

    c["opendate"] = None
    c["closedate"] = None

    if "Operating" in datestr:
        c["status"] = "Operating"
        if "Operating since" in datestr:
            if "?" in datestr and args.skipnodate:
                if args.verbose > 0:
                    print("--Skipping " + name + " at " + park + " (operating, '?' opening date)")
                return None
            elif " - " in datestr:
                c["opendate"] = dates[0]['datetime'] + " - " + dates[1]['datetime']
            elif "≤" in datestr:
                c["opendate"] = "≤ " + dates[0]['datetime']
            elif "≥" in datestr:
                c["opendate"] = "≥ " + dates[0]['datetime']
            else:
                c["opendate"] = dates[0]['datetime']
            if args.skipwrongyear and len(dates) > 0:
                if int(dates[0]['datetime'][:4]) > int(args.setyear):
                    if args.verbose > 0:
                        print("--Skipping " + name + " at " + park + " (opened after " + args.setyear + ")")
                    return None
        elif args.skipnodate:
            if args.verbose > 0:
                print("--Skipping " + name + " at " + park + " (operating, unknown opening date)")
            return None

    elif "Removed" in datestr:
        c["status"] = "Removed"
        if "Operated from" in datestr:

            # apologies for the spaghetti code; too many edge cases to check
            minopendate = None
            maxclosedate = None
            if "from ? to" in datestr:
                if args.skipnodate:
                    if args.verbose > 0:
                        print("--Skipping " + name + " at " + park + " (removed, '?' opening date)")
                    return None
                c["opendate"] = "?"
                if " - " in datestr:
                    c["closedate"] =  dates[0]['datetime'] + " - " + dates[1]['datetime']
                    maxclosedate = int(dates[1]['datetime'][:4])
                elif "≤" in datestr:
                    c["closedate"] = "≤ " + dates[0]['datetime']
                elif "≥" in datestr:
                    c["closedate"] = "≥ " + dates[0]['datetime']
                else:
                    c["closedate"] = dates[0]['datetime']
                if maxclosedate is None:
                    maxclosedate = int(dates[0]['datetime'][:4])
            elif "to ?" in datestr:
                c["closedate"] = "?"
                if " - " in datestr:
                    c["opendate"] = dates[0]['datetime'] + " - " + dates[1]['datetime']
                elif "≤" in datestr:
                    c["opendate"] = "≤ " + dates[0]['datetime']
                elif "≥" in datestr:
                    c["opendate"] = "≥ " + dates[0]['datetime']
                else:
                    c["opendate"] = dates[0]['datetime']
                minopendate = int(dates[0]['datetime'][:4])
            elif "from  -  to" in datestr:
                c["opendate"] = dates[0]['datetime'] + " - " + dates[1]['datetime']
                minopendate = int(dates[0]['datetime'][:4])
                if "to  - " in datestr:
                    c["closedate"] =  dates[2]['datetime'] + " - " + dates[3]['datetime']
                    maxclosedate = int(dates[3]['datetime'][:4])
                elif "≤" in datestr:
                    c["closedate"] = "≤ " + dates[2]['datetime']
                elif "≥" in datestr:
                    c["closedate"] = "≥ " + dates[2]['datetime']
                else:
                    c["closedate"] = dates[2]['datetime']
                if maxclosedate is None:
                    maxclosedate = int(dates[2]['datetime'][:4])
            elif "from ≤  to" in datestr:
                c["opendate"] = "≤ " + dates[0]['datetime']
                minopendate = int(dates[0]['datetime'][:4])
                if " - " in datestr:
                    c["closedate"] =  dates[1]['datetime'] + " - " + dates[2]['datetime']
                    maxclosedate = int(dates[2]['datetime'][:4])
                elif "≤" in datestr:
                    c["closedate"] = "≤ " + dates[1]['datetime']
                elif "≥" in datestr:
                    c["closedate"] = "≥ " + dates[1]['datetime']
                else:
                    c["closedate"] = dates[1]['datetime']
                if maxclosedate is None:
                    maxclosedate = int(dates[1]['datetime'][:4])
            elif "from ≥  to" in datestr:
                c["opendate"] = "≥ " + dates[0]['datetime']
                minopendate = int(dates[0]['datetime'][:4])
                if " - " in datestr:
                    c["closedate"] =  dates[1]['datetime'] + " - " + dates[2]['datetime']
                    maxclosedate = int(dates[2]['datetime'][:4])
                elif "≤" in datestr:
                    c["closedate"] = "≤ " + dates[1]['datetime']
                elif "≥" in datestr:
                    c["closedate"] = "≥ " + dates[1]['datetime']
                else:
                    c["closedate"] = dates[1]['datetime']
                if maxclosedate is None:
                    maxclosedate = int(dates[1]['datetime'][:4])
            elif "from  to" in datestr:
                c["opendate"] = dates[0]['datetime']
                minopendate = int(dates[0]['datetime'][:4])
                if " - " in datestr:
                    c["closedate"] =  dates[1]['datetime'] + " - " + dates[2]['datetime']
                    maxclosedate = int(dates[2]['datetime'][:4])
                elif "≤" in datestr:
                    c["closedate"] = "≤ " + dates[1]['datetime']
                elif "≥" in datestr:
                    c["closedate"] = "≥ " + dates[1]['datetime']
                else:
                    c["closedate"] = dates[1]['datetime']
                if maxclosedate is None:
                    maxclosedate = int(dates[1]['datetime'][:4])
            else:
                c["opendate"] = dates[0]['datetime']
                minopendate = int(dates[0]['datetime'][:4])
                if len(dates) > 1:
                    print("Something went wrong in the open/close date parsing...")

            # determine if the ride operated in the given year (if -y arg is used)
            if args.skipwrongyear:
                if minopendate is not None and minopendate > int(args.setyear):
                    if args.verbose > 0:
                        print("--Skipping " + name + " at " + park + " (opened after " + args.setyear + ")")
                    return None
                if maxclosedate is not None and maxclosedate < int(args.setyear):
                    if args.verbose > 0:
                        print("--Skipping " + name + " at " + park + " (removed before " + args.setyear + ")")
                    return None
                if minopendate is None:
                    if args.verbose > 0:
                        print("--Skipping " + name + " at " + park + " (removed, unknown opening year)")
                    return None
                if maxclosedate is None:
                    if args.verbose > 0:
                        print("--Skipping " + name + " at " + park + " (removed, unknown closing year)")
                    return None
        elif args.skipnodate:
            if args.verbose > 0:
                print("--Skipping " + name + " at " + park + " (removed, unknown opening date)")
            return None

    elif "SBNO" in datestr:
        c["status"] = "SBNO"
        if not csoup.find('table', attrs={'class':'objDemoBox'}): # for Orphan Rocker
            if args.skipnodate or args.skipwrongyear:
                if args.verbose > 0:
                    print("--Skipping " + name + " at " + park + " (SBNO, never operated)")
                return None
        else:
            for tr in csoup.find_all('table', attrs={'class':'objDemoBox'})[-1].find_all('tr'):
                if "Former status" in tr.text:
                    td = tr.find_all('td')[-1]
                    tdates = td.find_all('time')
                    tdtext = re.split('Operate|SBN', td.text)[1:] # hacky text operations incoming
                    minopendate = None
                    maxclosedate = None

                    # check first line of "Operated from" for closing date
                    if "d" is tdtext[0][0]:
                        if "to ?" in tdtext[0]:
                            c["closedate"] = "?"
                        elif "from ? to" in tdtext[0]:
                            if " - " in tdtext[0]:
                                c["closedate"] = tdates[0]['datetime'] + " - " + tdates[1]['datetime']
                                maxclosedate = int(tdates[1]['datetime'][:4])
                            else:
                                if "to ≤" in tdtext[0]:
                                    c["closedate"] = "≤ " + tdates[0]['datetime']
                                elif "to ≥" in tdtext[0]:
                                    c["closedate"] = "≥ " + tdates[0]['datetime']
                                else:
                                    c["closedate"] = tdates[0]['datetime']
                                maxclosedate = int(tdates[0]['datetime'][:4])
                        elif "to" in tdtext[0] and tdtext[0].count('-') > 1:
                            c["closedate"] = tdates[2]['datetime'] + " - " + tdates[3]['datetime']
                            maxclosedate = int(tdates[3]['datetime'][:4])
                        elif "to" in tdtext[0] and tdtext[0].count('-') > 0:
                            if " -  to" in tdtext[0]:
                                if "to ≤" in tdtext[0]:
                                    c["closedate"] = "≤ " + tdates[2]['datetime']
                                elif "to ≥" in tdtext[0]:
                                    c["closedate"] = "≥ " + tdates[2]['datetime']
                                else:
                                    c["closedate"] = tdates[2]['datetime']
                            else:
                                c["closedate"] = tdates[1]['datetime'] + " - " + tdates[2]['datetime']
                            maxclosedate = int(tdates[2]['datetime'][:4])
                        elif "to ≤" in tdtext[0]:
                            c["closedate"] = "≤ " + tdates[1]['datetime']
                            maxclosedate = int(tdates[1]['datetime'][:4])
                        elif "to ≥" in tdtext[0]:
                            c["closedate"] = "≥ " + tdates[1]['datetime']
                            maxclosedate = int(tdates[1]['datetime'][:4])
                        elif "to" in tdtext[0]:
                            c["closedate"] = tdates[1]['datetime']
                            maxclosedate = int(tdates[1]['datetime'][:4])

                    # check last line of "Operated from" for opening date
                    if "d" is tdtext[-1][0]:
                        if "from ? to" in tdtext[-1]:
                            if args.skipnodate:
                                if args.verbose > 0:
                                    print("--Skipping " + name + " at " + park + " (SBNO, '?' opening date)")
                                return None
                            c["opendate"] = "?"
                        elif "from  -  to" in tdtext[-1]:
                            if "to  - " in tdtext[-1]:
                                c["opendate"] = tdates[-4]['datetime'] + " - " + tdates[-3]['datetime']
                                minopendate = int(tdates[-4]['datetime'][:4])
                            elif "to ?" in tdtext[-1]:
                                c["opendate"] = tdates[-1]['datetime']
                                minopendate = int(tdates[-1]['datetime'][:4])
                            else:
                                c["opendate"] = tdates[-3]['datetime'] + " - " + tdates[-2]['datetime']
                                minopendate = int(tdates[-3]['datetime'][:4])
                        elif "from ≤" in tdtext[-1]:
                            if "to  - " in tdtext[-1]:
                                c["opendate"] = "≤ " + tdates[-3]['datetime']
                                minopendate = int(tdates[-3]['datetime'][:4])
                            elif "to ?" in tdtext[-1]:
                                c["opendate"] = tdates[-1]['datetime']
                                minopendate = int(tdates[-1]['datetime'][:4])
                            else:
                                c["opendate"] = "≤ " + tdates[-2]['datetime']
                                minopendate = int(tdates[-2]['datetime'][:4])
                        elif "from ≥" in tdtext[-1]:
                            if "to  - " in tdtext[-1]:
                                c["opendate"] = "≥ " + tdates[-3]['datetime']
                                minopendate = int(tdates[-3]['datetime'][:4])
                            elif "to ?" in tdtext[-1]:
                                c["opendate"] = tdates[-1]['datetime']
                                minopendate = int(tdates[-1]['datetime'][:4])
                            else:
                                c["opendate"] = "≥ " + tdates[-2]['datetime']
                                minopendate = int(tdates[-2]['datetime'][:4])
                        elif "to" in tdtext[-1]:
                            if "to  - " in tdtext[-1]:
                                c["opendate"] = tdates[-3]['datetime']
                                minopendate = int(tdates[-3]['datetime'][:4])
                            elif "to ?" in tdtext[-1]:
                                c["opendate"] = tdates[-1]['datetime']
                                minopendate = int(tdates[-1]['datetime'][:4])
                            else:
                                c["opendate"] = tdates[-2]['datetime']
                                minopendate = int(tdates[-2]['datetime'][:4])
                        else:
                            if "-" in tdtext[-1]:
                                c["opendate"] = tdates[-2]['datetime'] + " - " + tdates[-1]['datetime']
                                minopendate = int(tdates[-2]['datetime'][:4])
                            else:
                                c["opendate"] = tdates[-1]['datetime']
                                minopendate = int(tdates[-1]['datetime'][:4])

                    # determine if the ride operated in the given year (if -y arg is used)
                    if args.skipwrongyear:
                        if minopendate is not None and minopendate > int(args.setyear):
                            if args.verbose > 0:
                                print("--Skipping " + name + " at " + park + " (opened after " + args.setyear + ")")
                            return None
                        if maxclosedate is not None and maxclosedate < int(args.setyear):
                            if args.verbose > 0:
                                print("--Skipping " + name + " at " + park + " (SBNO before " + args.setyear + ")")
                            return None
                        if minopendate is None:
                            if args.verbose > 0:
                                print("--Skipping " + name + " at " + park + " (SBNO, unknown opening year)")
                            return None
                        if maxclosedate is None:
                            if args.verbose > 0:
                                print("--Skipping " + name + " at " + park + " (SBNO, unknown closing year)")
                            return None

                    break

    elif "In Storage" in datestr:
        c["status"] = "In Storage"
        return None # don't care about coasters in storage, for now
                    # very few are listed, see https://rcdb.com/r.htm?st=312&ot=2

    elif "Under Construction" in datestr:
        c["status"] = "Under Construction"
        return None # don't care about rides that have yet to be constructed

    # get thrill scale
    linkrow = feature.find('span', attrs={'class':'link_row'})
    if "Kiddie" in linkrow.text:
        if args.skipkiddie:
            if args.verbose > 0:
                print("--Skipping " + name + " at " + park + " (\"Kiddie\" designation)")
            return None
        c["scale"] = "Kiddie"
    elif "Family" in linkrow.text:
        c["scale"] = "Family"
    elif "Thrill" in linkrow.text:
        c["scale"] = "Thrill"
    elif "Extreme" in linkrow.text:
        c["scale"] = "Extreme"

    # get make and model
    makemodellist = feature.find_all('div', attrs={'class':'scroll'})
    if len(makemodellist) > 1:
        makemodel = makemodellist[1].text
        if "Make: " in makemodel:
            make = makemodel.split("Make: ", 1)[1]
            if "Model: " in makemodel:
                model = make.split("Model: ", 1)[1]
                make = make.split("Model: ", 1)[0]
                if " / " in model:
                    c["submodel"] = model.split(" / ", 1)[1]
                    model = model.split(" / ", 1)[0]
                c["model"] = model
            c["make"] = make

    # get track layout
    if "Track layout: " in feature.text:
        layout = feature.text.split("Track layout: ", 1)[1]
        if "Pictures" in layout:
            layout = layout.split("Pictures", 1)[0]
        elif "Videos" in layout:
            layout = layout.split("Videos", 1)[0]
        elif "Maps" in layout:
            layout = layout.split("Maps", 1)[0]
        if args.skipwackyworm and "Big Apple / Wacky Worm" in layout:
            if args.verbose > 0:
                print("--Skipping " + name + " at " + park + " (Big Apple / Wacky Worm)")
            return None
        if args.skipbutterfly and "U Shuttle" in layout and "scale" in c:
            if c["scale"] == "Kiddie" or c["scale"] == "Family":
                if args.verbose > 0:
                    print("--Skipping " + name + " at " + park + " (Butterfly)")
                return None
        c["layout"] = layout


    # get number of tracks; write new tracks to separate coaster entries
    stats = csoup.body.find('table', attrs={'id':'statTable'})
    numTracks = len(stats.find('tr').find_all('td'))
    c["tracks"] = str(numTracks)
    trackNames = []
    d = []
    if numTracks > 1 and args.combineTracks is False:
        if stats.find('th').text == "Name":
            for x in stats.find('tr').find_all('td'):
                trackNames.append(x.text)
        for i in range(numTracks):
            c["id"] = chr(ord('a')+i) + url.split("/")[-1].split(".")[0]
            if len(trackNames) > 1:
                c["name"] = "\"" + name + " (" + trackNames[i] + ")\""
            else:
                c["name"] = "\"" + name + " (" + chr(ord('a')+i) + ")\""
            d.append(c.copy())
    else:
        c["id"] = "r" + url.split("/")[-1].split(".")[0]

    # verbose print info per coaster
    if args.verbose > 1: # -vv = coaster name
        print(name, end="")
        if "altname" in c:
            print(" (AKA " + altname + ")", end="")
        print(", " + park, end="")
        if args.verbose > 2: # -vvv = coaster location (same line as coaster name)
            print(" - " + location, end="")
        print("")
        if args.verbose > 3: # -vvvv = coaster status + opening/closing dates + # of tracks
            if "status" in c:
                print("    Status: " + c["status"], end="")
                if "opendate" in c and c["opendate"] is not None:
                    if c["status"] is "Operating":
                        print(" since " + c["opendate"], end="")
                    elif c["status"] is "Removed" or "SBNO":
                        print(", Operated from " + c["opendate"], end="")
                    else:
                        print(", ", end="")
                    if "closedate" in c and c["closedate"] is not None:
                        print(" to " + c["closedate"], end="")
                print("")
                if numTracks > 1:
                    print("    Tracks: " + str(numTracks), end="")
                    if len(trackNames) > 1:
                        print(" - ", end="")
                        for i in range(len(trackNames)):
                            print(trackNames[i], end="")
                            if i < len(trackNames) - 1:
                                print(", ", end="")
                    print("")
            if args.verbose > 4: # -vvvvv = coaster scale + make/model + layout
                if "scale" in c:
                    print("    Scale:  " + c["scale"])
                if "make" in c:
                    print("    Make:   " + c["make"])
                if "model" in c:
                    if "submodel" in c:
                        print("    Model:  " + c["model"] + " / " + c["submodel"])
                    else:
                        print("    Model:  " + c["model"])
                if "layout" in c:
                    print("    Layout: " + c["layout"])

    if numTracks == 1 or args.combineTracks is True:
        return c
    else:
        return d

if __name__ == "__main__": # allows us to put main at the beginning
    main()

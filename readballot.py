#!/usr/bin/env python3

# =========================================================================
#  ElloCoaster reader for detailed blank ballots and regular voter ballots
#  Author: Grant Barker
#
#  Requires Python 3
# =========================================================================

from coaster import Coaster
from openpyxl import load_workbook

def blank_to_none(string):
    if string:
        return string
    else:
        return None

def read_detailed_ballot(filepath, id_col=7, name_col=2, park_col=4):
    coasters = {}
    wb = load_workbook(filepath)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2): 
        rcid = row[id_col].value
        url = None
        if "HYPERLINK" in row[id_col].value:
            rcid = row[id_col].value[:-2].split('"')[-1]
            url = row[id_col].value.split('"')[1]

        name = row[name_col].value
        park = row[park_col].value

        # silly exception for Gravity Group coasters listed under M&V on RCDB
        designer = blank_to_none(row[16].value)
        if designer == "Martin & Vleminckx":
            if name != "Coastersaurus" and park != "Legoland Florida":
                designer = "The Gravity Group, LLC"

        # silly exception for two notable Schwarzkopf coasters in Sweden
        # https://rcdb.com/1061.htm and https://rcdb.com/1062.htm
        if rcid == "r1061" or rcid == "r1062":
            designer = "Schwarzkopf"

        # silly exception for Powder Keg at Silver Dollar City
        if rcid == "r1903":
            designer = "S&S Sansei Technologies"

        c = Coaster(rcid,                         #rcid
                    name,                         #name
                    park,                         #park
                    url,                          #url
                    blank_to_none(row[3].value),  #altname
                    blank_to_none(row[5].value),  #country
                    blank_to_none(row[6].value),  #fullcity
                    blank_to_none(row[8].value),  #location
                    blank_to_none(row[9].value),  #state
                    blank_to_none(row[10].value), #city
                    blank_to_none(row[11].value), #status
                    blank_to_none(row[12].value), #opendate
                    blank_to_none(row[13].value), #closedate
                    blank_to_none(row[14].value), #rctype
                    blank_to_none(row[15].value), #scale
                    designer                    , #make
                    blank_to_none(row[17].value), #model
                    blank_to_none(row[18].value), #submodel
                    blank_to_none(row[19].value)) #tracks
        coasters[rcid] = c

    return coasters

def is_rcid(string):
    if isinstance(string, str):
        if "HYPERLINK" in string:
            return is_rcid(string[:-2].split('"')[-1])
        if len(string) < 7 and string[1:].isdecimal():
            return string
    return False

### ========================================================================
###  xls support copied/modified from these helpful users at StackOverflow
###  https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx
### ========================================================================
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import InvalidFileException

def open_xls_as_xlsx(filepath):
    # first open using xlrd
    book = xlrd.open_workbook(filepath)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1
### ========================================================================

def read_xlsx_ballot(wb, filepath):
    voter_rankings = {}
    ws = wb.worksheets[0]
    for row in ws.iter_rows():
        if isinstance(row[0].value, int) or isinstance(row[0].value, float):
            if row[0].value > 0 and row[1].value != "No":
                try:
                    rcid = row[7].value
                except IndexError:
                    rcid = None
                if not is_rcid(rcid):
                    rcid = None
                    for cell in row:
                        if is_rcid(cell.value):
                            rcid = cell.value
                    if rcid is None:
                        continue
                if "HYPERLINK" in rcid:
                    rcid = rcid[:-2].split('"')[-1]
                rank = row[0].value
                voter_rankings[rcid] = rank
    if len(voter_rankings) < 1:
        print("Error: " + filepath + " has misaligned/missing columns")
        return {}
    return voter_rankings

def isnum(numstring):
    if numstring.isdecimal():
        return True
    try:
        float(numstring)
        return True
    except ValueError:
        return False

import csv

def read_csv_ballot(filepath):
    voter_rankings = {}
    with open(filepath) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            if isnum(row[0]) and int(row[0]) > 0 and row[1] != "No":
                try:
                    rcid = row[7]
                except IndexError:
                    rcid = None
                if not is_rcid(rcid):
                    rcid = None
                    for cell in row:
                        if is_rcid(cell):
                            rcid = cell
                    if rcid is None:
                        continue
                rank = float(row[0])
                voter_rankings[rcid] = rank
    if len(voter_rankings) < 1:
        print("Error: " + filepath + " has misaligned/missing columns")
        return {}
    return voter_rankings

def read_voter_ballot(filepath, i=False):
    if filepath[-4:] == ".csv":
        return read_csv_ballot(filepath)
    if filepath[-4:] == ".xls":
        return read_xlsx_ballot(open_xls_as_xlsx(filepath), filepath)
    if filepath[-5:] == ".xlsx":
        return read_xlsx_ballot(load_workbook(filepath), filepath)
    print("Error: " + filepath + " is not a valid .csv/.xls/.xlsx file")
    return None

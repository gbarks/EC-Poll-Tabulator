#!/usr/bin/env python3

# ==========================================================
#  ElloCoaster poll tabulator
#  Author: Grant Barker
#  Contributions from Dave Wong, Jim Winslett
#
#  Requires Python 3
# ==========================================================

from __future__ import print_function

import sys

if sys.version_info[0] < 3:
    print("Program requires Python 3; running {0}.{1}".format(sys.version_info[0], sys.version_info[1]))
    sys.exit()

import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# nonessential local imports
useSpinner = True
try:
    from spinner import Spinner
except:
    useSpinner = False

# essential local imports
try:
    from coaster import Coaster
except:
    print('Could not find "coaster.py"; exiting...')
    sys.exit()

# global strings for parsing ballots
commentStr = "* "
blankUserField = "-Replace "
startLine = "! DO NOT CHANGE OR DELETE THIS LINE !"

# command line arguments
parser = argparse.ArgumentParser(description='Process Mitch Hawker-style coaster poll.')

parser.add_argument("-b", "--blankBallot", default="blankballot2019.txt",
                    help="specify blank ballot file")
parser.add_argument("-f", "--ballotFolder", default="ballots2019",
                    help="specify folder containing filled ballots")
parser.add_argument("-m", "--minRiders", type=int, default=10,
                    help="specify minimum number of riders for a coaster to rank")
parser.add_argument("-o", "--outfile", default="Poll Results.xlsx",
                    help="specify name of output .xlsx file")
parser.add_argument("-c", "--colorize", action="store_true",
                    help="color coaster designers in spreadsheet (requires -r)")
parser.add_argument("-d", "--designset", default="wood",
                    help="specify design/manufacturer dictionary (wood or steel)")
parser.add_argument("-i", "--includeExtraInfo", action="count", default=0,
                    help="include voter data/misc info; duplicate for more info")
parser.add_argument("-r", "--botherRCDB", action="store_true",
                    help="bother RCDB to grab metadata from links in blankBallot")
parser.add_argument("-v", "--verbose", action="count", default=0,
                    help="print data as it's processed; duplicate for more info")

args = parser.parse_args()

if not os.path.isfile(args.blankBallot):
    print('Blank ballot source "{0}" is not a file; exiting...'.format(args.blankBallot))
    sys.exit()

if not os.path.isdir(args.ballotFolder) or len(os.listdir(args.ballotFolder)) < 1:
    print('Ballot folder "{0}" does not exist or is empty; exiting...'.format(args.ballotFolder))
    sys.exit()

if args.outfile[-5:] != ".xlsx":
    args.outfile += ".xlsx"

# import the correct set of designers/manufacturers
if args.designset.lower() == "wood":
    try:
        from wood import designers
    except:
        print('Could not find "wood.py"; exiting...')
        sys.exit()
elif args.designset.lower() == "steel":
    print("Steel poll functionality hasn't been implemented yet; sorry...")
    sys.exit()
else:
    print("Wood and steel are the only valid design/manufacturer dictionaries; exiting...")
    sys.exit()

# colorizing coasters by designer requires fetching RCDB data
if args.colorize and not args.botherRCDB:
    args.colorize = False



# ==================================================
#  onto main()!
# ==================================================

def main():
    # create Excel workbook
    xlout = Workbook()
    xlout.active.title = "Coaster Masterlist"

    # preferred fixed-width font
    menlo = Font(name="Menlo")

    # list of tuples of the form (fullCoasterName, abbreviatedCoasterName)
    coasterDict = getCoasterDict(xlout.active, menlo)

    # create color key for designers
    if args.colorize:
        coasterdesignerws = xlout.create_sheet("Coaster Designer Color Key")
        i = 1
        for designer in sorted(designers.keys()):
            if designer != "" and designer != "Other Known Manufacturer":
                coasterdesignerws.append([designer])
                coasterdesignerws.cell(row=i, column=1).fill = designers[designer]
                i += 1
        if "Other Known Manufacturer" in designers.keys():
            coasterdesignerws.append(["Other Known Manufacturer"])
            coasterdesignerws.cell(row=i, column=1).fill = designers["Other Known Manufacturer"]
            i += 1
        if "" in designers.keys():
            coasterdesignerws.append(["Other [Unknown]"])
            coasterdesignerws.cell(row=i, column=1).fill = designers[""]
        coasterdesignerws.column_dimensions['A'].width = 30.83

    # for each pair of coasters, a list of numbers of the form [wins, losses, ties, winPercent]
    winLossMatrix = createMatrix(coasterDict)

    processAllBallots(xlout, coasterDict, winLossMatrix)

    calculateResults(coasterDict, winLossMatrix)

    # sorted lists of tuples of the form (rankedCoaster, relevantNumbers)
    finalResults, finalPairs = sortedLists(coasterDict, winLossMatrix)

    # write worksheets related to finalResults, finalPairs, and winLossMatrix
    printToFile(xlout, finalResults, finalPairs, winLossMatrix, coasterDict, menlo, designers)

    # save the Excel file
    print("Saving...", end=" ")
    if useSpinner:
        spinner = Spinner()
        spinner.start()
    xlout.save(args.outfile)
    if useSpinner:
        spinner.stop()
    print('output saved to "{0}".'.format(args.outfile))



# ==================================================
#  function for getting manufacturer's color
# ==================================================

def colorizeRow(worksheet, rowNum, colList, coasterDict, coaster, colorDict):
    if args.colorize:
        if coasterDict[coaster].designer:
            if coasterDict[coaster].designer in colorDict.keys():
                for l in colList:
                    worksheet.cell(row=rowNum, column=l).fill = colorDict[coasterDict[coaster].designer]
            else:
                for l in colList:
                    worksheet.cell(row=rowNum, column=l).fill = colorDict["Other Known Manufacturer"]
        else:
            for l in colList:
                    worksheet.cell(row=rowNum, column=l).fill = colorDict[""]



# ==================================================
#  populate dictionary of coasters in the poll
# ==================================================

def getCoasterDict(masterlistws, preferredFixedWidthFont):
    if not args.botherRCDB:
        print("Creating list of every coaster on the ballot...", end=" ")
        if useSpinner:
            spinner = Spinner()
            spinner.start()

    # set up Coaster Masterlist worksheet
    headerRow = ["Full Coaster ID", "Abbrev.", "Name", "Park", "Loc."]
    if args.botherRCDB:
        headerRow.extend(["RCDB Link", "Designer/Manufacturer", "Year"])
    masterlistws.append(headerRow)
    masterlistws.column_dimensions['A'].width = 45.83
    masterlistws.column_dimensions['B'].width = 12.83
    masterlistws.column_dimensions['C'].width = 25.83
    masterlistws.column_dimensions['D'].width = 25.83
    masterlistws.column_dimensions['E'].width = 6.83
    masterlistws['B1'].font = preferredFixedWidthFont
    masterlistws['E1'].font = preferredFixedWidthFont
    if args.botherRCDB:
        masterlistws.column_dimensions['F'].width = 16.83
        masterlistws.column_dimensions['G'].width = 25.83
        masterlistws.column_dimensions['H'].width = 4.83

    coasterDict = {} # return value

    #open the blank ballot file
    with open(args.blankBallot) as f:
        lineNum = 0
        startProcessing = False

        # begin going through the blank ballot line by line
        for line in f:

            sline = line.strip() # strip whitespace from start and end of line
            lineNum += 1

            # skip down the file to the coasters
            if startProcessing == False and sline == startLine:
                startProcessing = True

            # add the coasters to coasterDict and the masterlist worksheet
            elif startProcessing == True:

                if commentStr in sline: # skip comment lines (begin with "* ")
                    continue

                elif sline == "": # skip blank lines
                    continue

                else:
                    # break the line into its components: rank, full coaster name, abbreviation
                    words = [x.strip() for x in sline.split(',')]

                    # make sure there are at least 3 'words' in each line (rank, fullName, abbrName)
                    if len(words) < 3:
                        print("Error in {0}, Line {1}: {2}".format(args.blankBallot, lineNum, line))

                    else:
                        if len(words) > 3 and args.botherRCDB:
                            c = Coaster(words[1], words[2], words[3], designers.keys())
                        else:
                            c = Coaster(words[1], words[2])

                        # list of strings that will form a row in the spreadsheet
                        rowVals = [c.uniqueID, c.abbr, c.name, c.park, c.location]

                        # add RCDB-pulled info to spreadsheet row and print it
                        if len(words) > 3 and args.botherRCDB:
                            rowVals.append('=HYPERLINK("{0}", "{1}")'.format(c.rcdb, c.rcdb[8:]))
                            rowVals.extend([c.designer, c.year])
                            print("{0},   \t{1},\t{2}".format(c.abbr, c.year, c.designer))

                        # append the row values and set styles
                        masterlistws.append(rowVals)
                        masterlistws.cell(row=len(coasterDict)+1, column=5).font = preferredFixedWidthFont
                        masterlistws.cell(row=len(coasterDict)+1, column=2).font = preferredFixedWidthFont
                        if c.rcdb:
                            masterlistws.cell(row=len(coasterDict)+1, column=6).style = "Hyperlink"

                        # add the coaster to the dictionary of coasters on the ballot
                        coasterDict[c.uniqueID] = c

                        colorizeRow(masterlistws, len(coasterDict)+1, [1,2,7], coasterDict, c.uniqueID, designers)

    masterlistws.freeze_panes = masterlistws['A2']
    if useSpinner and not args.botherRCDB:
        spinner.stop()
    print("{0} coasters on the ballot.".format(len(coasterDict)))
    return coasterDict



# ==================================================
#  import filepaths of ballots
# ==================================================

def getBallotFilepaths():
    print("Getting the filepaths of submitted ballots...", end=" ")
    if useSpinner:
        spinner = Spinner()
        spinner.start()

    ballotList = []
    for file in os.listdir(args.ballotFolder):
        if file.endswith(".txt"):
            ballotList.append(os.path.join(args.ballotFolder, file))

    if useSpinner:
        spinner.stop()
    print("{0} ballots submitted.".format(len(ballotList)))
    return ballotList



# ==================================================
#  create win/loss matrix
# ==================================================

def createMatrix(coasterDict):
    print("Creating the win/loss matrix...", end=" ")
    if useSpinner:
        spinner = Spinner()
        spinner.start()

    winLossMatrix = {}
    for coasterA in coasterDict.keys():
        for coasterB in coasterDict.keys():

            # can't compare a coaster to itself
            if coasterA != coasterB:
                winLossMatrix[coasterA, coasterB] = {}
                winLossMatrix[coasterA, coasterB]["Wins"] = 0
                winLossMatrix[coasterA, coasterB]["Losses"] = 0
                winLossMatrix[coasterA, coasterB]["Ties"] = 0
                winLossMatrix[coasterA, coasterB]["Win Percentage"] = 0.0
                winLossMatrix[coasterA, coasterB]["Pairwise Rank"] = 0

    if useSpinner:
        spinner.stop()
    print("{0} pairings.".format(len(winLossMatrix)))
    return winLossMatrix



# ================================================================
#  read a ballot (just ONE ballot)
#
#  you need a loop to call this function for each ballot filename
# ================================================================

def processBallot(filepath, coasterDict, winLossMatrix):
    filename = os.path.basename(filepath)
    if args.verbose > 0:
        print("Processing ballot: {0}".format(filename))

    voterInfo = [filename, "", "", "", "", ""] # return item 1
    coasterAndRank = {} # return item 2
    creditNum = 0
    error = False

    # open the ballot file
    with open(filepath) as f:
        infoField = 1
        lineNum = 0
        startProcessing = False

        for line in f:
            sline = line.strip()
            lineNum += 1

            # begin at top of ballot and get the voter's info first
            if startProcessing == False and infoField <= 5 and not commentStr in sline and len(sline) != 0:

                # if the line begins with "-Replace" then record a non-answer
                if blankUserField in sline:
                    voterInfo[infoField] = ""
                    infoField += 1
                elif not startLine in sline:
                    voterInfo[infoField] = sline.strip('-').strip()
                    infoField += 1

            # skip down the file to the coasters
            if startProcessing == False and sline == startLine:
                startProcessing = True

            elif startProcessing == True:

                # break the line into its components: rank, name
                words = [x.strip() for x in sline.split(',')]

                if commentStr in sline: # skip comment lines (begin with "* ")
                    continue

                elif sline == "": # skip blank lines
                    continue

                # make sure there are at least 2 'words' in each line
                elif len(words) < 2:
                    if args.verbose == 0:
                        print("Processing ballot: {0}".format(filename))
                    print("Error in {0}, Line {1}: {2}".format(args.blankBallot, lineNum, line))

                # make sure the ranking is a number
                elif not words[0].isdigit():
                    if args.verbose == 0:
                        print("Processing ballot: {0}".format(filename))
                    print("Error in reading {0}, Line {1}: Rank must be an int.".format(filename, lineNum))
                    error = True

                else:
                    coasterName = words[1]
                    coasterRank = int(words[0])

                    # skip coasters ranked zero or less (those weren't ridden)
                    if coasterRank <= 0:
                        continue

                    # check to make sure the coaster on the ballot is legit
                    if coasterName  in coasterDict.keys():
                        creditNum += 1
                        coasterDict[coasterName].riders += 1

                        # add this voter's ranking of the coaster
                        coasterAndRank[coasterName] = coasterRank

                    else: # it's not a legit coaster!
                        if args.verbose == 0:
                            print("Processing ballot: {0}".format(filename))
                        print("Error in reading {0}, Line {1}: Unknown coaster {2}".format(filename, lineNum, coasterName))
                        error = True

    # don't tally the ballot if there were any errors, don't return voter info
    if error:
        if args.verbose == 0:
            print("Processing ballot: {0}".format(filename))
        print("Error encountered. File {0} not added.".format(filename))
        return [], {}

    # cycle through each pair of coasters this voter ranked
    for coasterA in coasterAndRank.keys():
        for coasterB in coasterAndRank.keys():

            # can't compare a coaster to itself
            if coasterA != coasterB:

                # if the coasters have the same ranking, call it a tie
                if coasterAndRank[coasterA] == coasterAndRank[coasterB]:
                    winLossMatrix[coasterA, coasterB]["Ties"] += 1
                    coasterDict[coasterA].totalTies += 1

                # if coasterA outranks coasterB (the rank's number is lower), call it a win for coasterA
                elif coasterAndRank[coasterA] < coasterAndRank[coasterB]:
                    winLossMatrix[coasterA, coasterB]["Wins"] += 1
                    coasterDict[coasterA].totalWins += 1

                # if not a tie nor a win, it must be a loss
                else:
                    winLossMatrix[coasterA, coasterB]["Losses"] += 1
                    coasterDict[coasterA].totalLosses += 1

    if args.verbose > 0:
        print(" ->", end=" ")

        for i in range(1,len(voterInfo)):
            if voterInfo[i] != "":
                print("{0},".format(voterInfo[i]), end=" ")

        print("CC: {0}".format(creditNum))

    voterInfo.append(creditNum)

    return voterInfo, coasterAndRank



# ==================================================
#  read all ballots and mark spreadsheets
# ==================================================

def processAllBallots(xl, coasterDict, winLossMatrix):

    # include spreadsheet containing identifying voter info, if requested
    if args.includeExtraInfo > 0:
        voterinfows = xl.create_sheet("Voter Info (SENSITIVE)")
        voterinfows.append(["Ballot Filename","Name","Email","City","State/Province","Country","Coasters Ridden"])
        voterinfows.column_dimensions['A'].width = 24.83
        voterinfows.column_dimensions['B'].width = 16.83
        voterinfows.column_dimensions['C'].width = 24.83
        for col in ['D','E','F','G']:
            voterinfows.column_dimensions[col].width = 12.83

        # include spreadsheet containing individual ballots, if requested
        if args.includeExtraInfo > 1:

            # includes each coaster's rank in a separate column
            ballotws1 = xl.create_sheet("Ballots with Ranks (SENSITIVE)")
            headerRow = ["Ballot Filename"]
            for i in range(0, len(coasterDict)):
                headerRow.extend(["Rank","Coaster"])
            ballotws1.append(headerRow)
            ballotws1.column_dimensions['A'].width = 24.83
            for i in range(0, len(coasterDict)):
                col1 = (i * 2) + 2
                col2 = col1 + 1
                ballotws1.column_dimensions[get_column_letter(col1)].width = 4.83
                ballotws1.column_dimensions[get_column_letter(col2)].width = 45.83

            # doesn't include rank data; assumes no coasters are ranked the same
            ballotws2 = xl.create_sheet("Ballots Imprecise (SENSITIVE)")
            headerRow = ["Ballot Filename"]
            for i in range(0, len(coasterDict)):
                headerRow.append("Coaster [Rank {0}]".format(i+1))
            ballotws2.append(headerRow)
            ballotws2.column_dimensions['A'].width = 24.83
            for i in range(0, len(coasterDict)):
                ballotws2.column_dimensions[get_column_letter(i+2)].width = 45.83

    # loop over ballots, processing each and saving requested info
    for filepath in getBallotFilepaths():
        voterInfo, ballotRanks = processBallot(filepath, coasterDict, winLossMatrix)
        if args.includeExtraInfo > 0 and voterInfo:
            voterinfows.append(voterInfo)
            if args.includeExtraInfo > 1 and ballotRanks:
                rowVals1 = [voterInfo[0]]
                rowVals2 = [voterInfo[0]]
                for coasterAndRank in sorted(ballotRanks.items(), key=lambda x: x[1]):
                    if args.verbose > 2:
                        print("{0}.\t{1}".format(coasterAndRank[1], coasterAndRank[0]))
                    rowVals1.extend([coasterAndRank[1], coasterAndRank[0]])
                    rowVals2.append(coasterAndRank[0])
                ballotws1.append(rowVals1)
                ballotws2.append(rowVals2)
    if args.includeExtraInfo > 0:
        voterinfows.freeze_panes = voterinfows['A2']
        if args.includeExtraInfo > 1:
            ballotws1.freeze_panes = ballotws1['B2']
            ballotws2.freeze_panes = ballotws2['B2']



# ========================================================
#  calculate results
#
#  no need to loop through this, since it calculates with
#    numbers gathered when the ballots were processed
# ========================================================

def calculateResults(coasterDict, winLossMatrix):
    print("Calculating results...", end=" ")
    if useSpinner:
        spinner = Spinner()
        spinner.start()

    if args.verbose > 0:
        print("")

    # iterate through all the pairs in the matrix
    for coasterA in coasterDict.keys():
        for coasterB in coasterDict.keys():

            # can't compare a coaster to itself
            if coasterA != coasterB:
                pairWins = winLossMatrix[coasterA, coasterB]["Wins"]
                pairLoss = winLossMatrix[coasterA, coasterB]["Losses"]
                pairTies = winLossMatrix[coasterA, coasterB]["Ties"]
                pairContests = pairWins + pairLoss + pairTies

                if pairContests > 0:
                    winLossMatrix[coasterA, coasterB]["Win Percentage"] = (((pairWins + float(pairTies / 2)) / pairContests)) * 100
                    
                    if pairWins == pairLoss:
                        coasterDict[coasterA].pairwiseTies += 1
                    elif pairWins > pairLoss:
                        coasterDict[coasterA].pairwiseWins += 1
                    else:
                        coasterDict[coasterA].pairwiseLosses += 1

                    # only print pairwise results with '-vvvv' flag
                    if args.verbose > 3:
                        print("{0},{1},\tWins: {2},\tTies: {3},\t#Con: {4},\tWin%: {5}".format(
                            coasterDict[coasterA].abbr, coasterDict[coasterB].abbr,
                            pairWins, pairTies, pairContests, winLossMatrix[coasterA, coasterB]["Win Percentage"]))

    for x in coasterDict.keys():
        totalWins = coasterDict[x].totalWins
        totalLoss = coasterDict[x].totalLosses
        totalTies = coasterDict[x].totalTies
        totalContests = totalWins + totalLoss + totalTies

        pairWins = coasterDict[x].pairwiseWins
        pairLoss = coasterDict[x].pairwiseLosses
        pairTies = coasterDict[x].pairwiseTies
        pairContests = pairWins + pairLoss + pairTies

        if  totalContests > 0:
            coasterDict[x].totalWinPercentage = ((totalWins + float(totalTies/2)) / totalContests) * 100
            coasterDict[x].pairwiseWinPercentage = ((pairWins + float(pairTies/2)) / pairContests) * 100

            # print singular results with just a '-v' flag
            if args.verbose > 0:
                print("{0},\tWins:{1},{2}\tTies:{3},{4}\t#Con:{5},{6}\tWin%: {7}, \tPairWin%: {8}".format(
                    coasterDict[x].abbr, totalWins, pairWins, totalTies, pairTies, totalContests, pairContests,
                    round(coasterDict[x].totalWinPercentage, 3), round(coasterDict[x].pairwiseWinPercentage, 3)))

    if useSpinner:
        spinner.stop()
    print(" ")



# ==================================================
#  add to "Tied Coasters" variable in coasterDict
# ==================================================

def markTies(coasterDict, winLossMatrix, tiedCoasters):
    for coasterA in tiedCoasters:
        coastersTiedWithA = []
        for coasterB in tiedCoasters:
            if coasterA != coasterB:
                coastersTiedWithA.append(coasterB)
        coasterDict[coasterA].tiedCoasters = coastersTiedWithA

    # print Mitch Hawker-style pairwise matchups between tied coasters with '-v' flag
    if args.verbose > 0:
        print("  ===Tied===", end="\t")
        for coaster in tiedCoasters:
            print(" {0} ".format(coasterDict[coasterB].abbr), end="\t")
        print("")
        for coasterA in tiedCoasters:
            print("  {0}".format(coasterDict[coasterA].abbr), end="\t")
            for coasterB in tiedCoasters:
                cellStr = " "
                if coasterA != coasterB:
                    if winLossMatrix[coasterA, coasterB]["Wins"] > winLossMatrix[coasterA, coasterB]["Losses"]:
                        cellStr += "W "
                    elif winLossMatrix[coasterA, coasterB]["Wins"] < winLossMatrix[coasterA, coasterB]["Losses"]:
                        cellStr += "L "
                    else:
                        cellStr += "T "
                    cellStr += str(winLossMatrix[coasterA, coasterB]["Wins"]) + "-"
                    cellStr += str(winLossMatrix[coasterA, coasterB]["Losses"]) + "-"
                    cellStr += str(winLossMatrix[coasterA, coasterB]["Ties"])
                else:
                    cellStr += "       "
                print("{0}".format(cellStr), end="\t")
            print("")



# ==================================================
#  create sorted list of coasters by win pct and
#    sorted list of coasters by pairwise win pct
# ==================================================

def sortedLists(coasterDict, winLossMatrix):
    print("Sorting the results...", end=" ")
    if useSpinner:
        spinner = Spinner()
        spinner.start()

    results = []
    pairPercents = []

    # iterate through coasterDict by coasters
    for coasterName in coasterDict.keys():
        if int(coasterDict[coasterName].riders) >= int(args.minRiders):
            results.append((coasterName,
                            coasterDict[coasterName].totalWinPercentage,
                            coasterDict[coasterName].pairwiseWinPercentage))

    # iterate through winLossMatrix by coaster pairings
    for coasterPair in winLossMatrix.keys():      
        pairPercents.append((coasterPair, winLossMatrix[coasterPair]["Win Percentage"]))

    # sort lists by win percentages
    sortedResults = sorted(results, key=lambda x: x[1], reverse=True)
    sortedPairs = sorted(pairPercents, key=lambda x: x[1], reverse=True)

    if args.verbose > 0:
        print("")

    # determine rankings including ties for overall rank
    overallRank = 0
    curRank = 0
    curValue = 0.0
    tiedCoasters = []
    for x in sortedResults:
        overallRank += 1
        if x[1] != curValue:
            if len(tiedCoasters) > 1: # do stuff on complete list of tied coasters
                markTies(coasterDict, winLossMatrix, tiedCoasters)
            curRank = overallRank
            curValue = x[1]
            tiedCoasters = []
        tiedCoasters.append(x[0])
        coasterDict[x[0]].overallRank = curRank
        if args.verbose > 0:
            print("Rank: {0},\tVal: {1},  \tCoaster: {2}".format(coasterDict[x[0]].overallRank, x[1], x[0]))
    if len(tiedCoasters) > 1: # in case last few coasters were tied
        markTies(coasterDict, winLossMatrix, tiedCoasters)

    # determine rankings including ties for pairwise ranks
    overallRank = 0
    curRank = 0
    curValue = 0.0
    for x in sortedPairs:
        overallRank += 1
        if x[1] != curValue:
            curRank = overallRank
            curValue = x[1]
        winLossMatrix[x[0][0], x[0][1]]["Pairwise Rank"] = curRank

    if useSpinner:
        spinner.stop()
    print(" ")

    return sortedResults, sortedPairs



# ==================================================
#  print everything to a file
# ==================================================

def printToFile(xl, results, pairs, winLossMatrix, coasterDict, preferredFixedWidthFont, manuColors):
    print("Writing the results...", end=" ")
    if useSpinner:
        spinner = Spinner()
        spinner.start()

    # create and write primary results worksheet
    resultws = xl.create_sheet("Ranked Results")
    headerRow = ["Rank","Coaster","Total Win Percentage","Pairwise Win Percentage",
                 "Total Wins","Total Losses","Total Ties","Pair Wins","Pair Losses",
                 "Pair Ties","Number of Riders"]
    if args.botherRCDB:
        headerRow.extend(["Designer/Manufacturer", "Year"])
    resultws.append(headerRow)
    resultws.column_dimensions['A'].width = 4.83
    resultws.column_dimensions['B'].width = 45.83
    resultws.column_dimensions['C'].width = 16.83
    resultws.column_dimensions['D'].width = 18.83
    resultws.column_dimensions['E'].width = 8.83
    resultws.column_dimensions['F'].width = 9.83
    resultws.column_dimensions['G'].width = 7.83
    resultws.column_dimensions['H'].width = 8.33
    resultws.column_dimensions['I'].width = 9.33
    resultws.column_dimensions['J'].width = 7.33
    resultws.column_dimensions['K'].width = 13.83
    if args.botherRCDB:
        resultws.column_dimensions['L'].width = 23.83
        resultws.column_dimensions['M'].width = 8.83
    i = 2
    for x in results:
        resultws.append([coasterDict[x[0]].overallRank, x[0],
                         coasterDict[x[0]].totalWinPercentage,
                         coasterDict[x[0]].pairwiseWinPercentage,
                         coasterDict[x[0]].totalWins,
                         coasterDict[x[0]].totalLosses,
                         coasterDict[x[0]].totalTies,
                         coasterDict[x[0]].pairwiseWins,
                         coasterDict[x[0]].pairwiseLosses,
                         coasterDict[x[0]].pairwiseTies,
                         coasterDict[x[0]].riders,
                         coasterDict[x[0]].designer,
                         coasterDict[x[0]].year])
        colorizeRow(resultws, i, [2,12], coasterDict, x[0], manuColors)
        i += 1
    resultws.freeze_panes = resultws['A2']

    # append coasters that weren't ranked to the bottom of results worksheet
    for x in coasterDict.keys():
        if x not in [y[0] for y in results] and coasterDict[x].riders > 0:
            resultws.append(["N/A", x,
                             "Insufficient Riders, {0}".format(coasterDict[x].totalWinPercentage),
                             "Insufficient Riders, {0}".format(coasterDict[x].pairwiseWinPercentage),
                             coasterDict[x].totalWins,
                             coasterDict[x].totalLosses,
                             coasterDict[x].totalTies,
                             coasterDict[x].pairwiseWins,
                             coasterDict[x].pairwiseLosses,
                             coasterDict[x].pairwiseTies,
                             coasterDict[x].riders,
                             coasterDict[x].designer,
                             coasterDict[x].year])
            colorizeRow(resultws, i, [2,12], coasterDict, x, manuColors)
            i += 1

    # append coasters that weren't ridden to the bottom of results worksheet
    for x in coasterDict.keys():
        if x not in [y[0] for y in results] and coasterDict[x].riders == 0:
            resultws.append(["N/A", x, "No Riders", "No Riders",
                             coasterDict[x].totalWins,
                             coasterDict[x].totalLosses,
                             coasterDict[x].totalTies,
                             coasterDict[x].pairwiseWins,
                             coasterDict[x].pairwiseLosses,
                             coasterDict[x].pairwiseTies,
                             coasterDict[x].riders,
                             coasterDict[x].designer,
                             coasterDict[x].year])
            colorizeRow(resultws, i, [2,12], coasterDict, x, manuColors)
            i += 1

    # create and write pairwise result worksheet
    pairws = xl.create_sheet("Ranked Pairs")
    pairws.append(["Rank","Primary Coaster","Rival Coaster","Win Percentage","Wins","Losses","Ties"])
    pairws.column_dimensions['A'].width = 4.83
    pairws.column_dimensions['B'].width = 45.83
    pairws.column_dimensions['C'].width = 45.83
    pairws.column_dimensions['D'].width = 12.83
    pairws.column_dimensions['E'].width = 4.5
    pairws.column_dimensions['F'].width = 5.5
    pairws.column_dimensions['G'].width = 3.83
    i = 2
    for x in pairs:
        pairws.append([winLossMatrix[x[0][0], x[0][1]]["Pairwise Rank"], x[0][0], x[0][1],
                       winLossMatrix[x[0][0], x[0][1]]["Win Percentage"],
                       winLossMatrix[x[0][0], x[0][1]]["Wins"],
                       winLossMatrix[x[0][0], x[0][1]]["Losses"],
                       winLossMatrix[x[0][0], x[0][1]]["Ties"]])
        colorizeRow(pairws, i, [2], coasterDict, x[0][0], manuColors)
        colorizeRow(pairws, i, [3], coasterDict, x[0][1], manuColors)
        i += 1
    pairws.freeze_panes = pairws['A2']

    # create and write Mitch Hawker-style mutual rider comparison worksheet
    hawkerWLTws = xl.create_sheet("Coaster vs Coaster Win-Loss-Tie")
    headerRow = ["Rank",""]
    for coaster in results:
        headerRow.append(coasterDict[coaster[0]].abbr)
    hawkerWLTws.append(headerRow)
    hawkerWLTws.column_dimensions['A'].width = 4.83
    hawkerWLTws.column_dimensions['B'].width = 45.83
    for col in range(3, len(results)+3):
        hawkerWLTws.column_dimensions[get_column_letter(col)].width = 12.83
        colorizeRow(hawkerWLTws, 1, [col], coasterDict, results[col-3][0], manuColors)
    for i in range(0, len(results)):
        resultRow = [coasterDict[results[i][0]].overallRank, results[i][0]]
        winCount = 0
        loseCount = 0
        tieCount = 0
        for j in range(0, len(results)):
            coasterA = results[i][0]
            coasterB = results[j][0]
            cellStr = ""
            if coasterA != coasterB:
                if winLossMatrix[coasterA, coasterB]["Wins"] > winLossMatrix[coasterA, coasterB]["Losses"]:
                    cellStr += "W "
                    winCount += 1
                elif winLossMatrix[coasterA, coasterB]["Wins"] < winLossMatrix[coasterA, coasterB]["Losses"]:
                    cellStr += "L "
                    loseCount += 1
                else:
                    cellStr += "T "
                    tieCount += 1
                cellStr += str(winLossMatrix[coasterA, coasterB]["Wins"]) + "-"
                cellStr += str(winLossMatrix[coasterA, coasterB]["Losses"]) + "-"
                cellStr += str(winLossMatrix[coasterA, coasterB]["Ties"])
            resultRow.append(cellStr)
        hawkerPct = ((winCount + (tieCount/float(2))/float(len(results)-1))* 100)
        resultRow.append(hawkerPct)
        hawkerWLTws.append(resultRow)
        colorizeRow(hawkerWLTws, i+2, [2], coasterDict, results[i][0], manuColors)
    hawkerWLTws.freeze_panes = hawkerWLTws['C2']
    for col in hawkerWLTws.iter_cols(min_col=3):
        for cell in col:
            cell.font = preferredFixedWidthFont

    # create and write Mitch Hawker-style mutual rider comparison worksheet sorted by Pairwise Win Percentage
    resortedResults = sorted(results, key=lambda x: x[2], reverse=True)
    hawkerWLT2 = xl.create_sheet("CvC Win-Loss-Tie by PairWin%")
    headerRow = ["Rank",""]
    if args.verbose > 0:
        print(" ")
    for x in resortedResults:
        headerRow.append(coasterDict[x[0]].abbr)
        if args.verbose > 0:
            print("Rank: {0},\tVal: {1},  \tCoaster: {2}".format(coasterDict[x[0]].overallRank, x[2], x[0]))
    hawkerWLT2.append(headerRow)
    hawkerWLT2.column_dimensions['A'].width = 4.83
    hawkerWLT2.column_dimensions['B'].width = 45.83
    for col in range(3, len(resortedResults)+3):
        hawkerWLT2.column_dimensions[get_column_letter(col)].width = 12.83
        colorizeRow(hawkerWLT2, 1, [col], coasterDict, resortedResults[col-3][0], manuColors)
    for i in range(0, len(resortedResults)):
        resultRow = [coasterDict[resortedResults[i][0]].overallRank, resortedResults[i][0]]
        for j in range(0, len(resortedResults)):
            coasterA = resortedResults[i][0]
            coasterB = resortedResults[j][0]
            cellStr = ""
            if coasterA != coasterB:
                if winLossMatrix[coasterA, coasterB]["Wins"] > winLossMatrix[coasterA, coasterB]["Losses"]:
                    cellStr += "W "
                elif winLossMatrix[coasterA, coasterB]["Wins"] < winLossMatrix[coasterA, coasterB]["Losses"]:
                    cellStr += "L "
                else:
                    cellStr += "T "
                cellStr += str(winLossMatrix[coasterA, coasterB]["Wins"]) + "-"
                cellStr += str(winLossMatrix[coasterA, coasterB]["Losses"]) + "-"
                cellStr += str(winLossMatrix[coasterA, coasterB]["Ties"])
            resultRow.append(cellStr)
        hawkerWLT2.append(resultRow)
        colorizeRow(hawkerWLT2, i+2, [2], coasterDict, resortedResults[i][0], manuColors)
    hawkerWLT2.freeze_panes = hawkerWLT2['C2']
    for col in hawkerWLT2.iter_cols(min_col=3):
        for cell in col:
            cell.font = preferredFixedWidthFont

    # create and write sheet to compare where coasters would have ranked if sorted by PairWin%
    comparisonws = xl.create_sheet("TotalWin% vs PairWin% Rankings")
    comparisonws.append(["Coaster","TotalWin% Rank","PairWin% Rank","Difference"])
    comparisonws.column_dimensions['A'].width = 45.83
    comparisonws.column_dimensions['B'].width = 12.83
    comparisonws.column_dimensions['C'].width = 12.83
    for i in range(0, len(resortedResults)):
        coaster = resortedResults[i][0]
        oldRank = coasterDict[coaster].overallRank
        newRank = i+1
        diff = oldRank - newRank
        if diff == 0:
            comparisonws.append([coaster, oldRank, newRank, ""])
        else:
            comparisonws.append([coaster, oldRank, newRank, diff])
            if diff <= -16:
                diffColor = "ff0000" # maximum red
            elif diff < 0:
                diffColor = "ff{0}{0}".format(hex(256 + 16 * diff)[2:]) # gradual red
            elif diff < 16:
                diffColor = "{0}ff{0}".format(hex(256 - 16 * diff)[2:]) # gradual green
            else:
                diffColor = "00ff00" # maximum green
            comparisonws.cell(row=i+2, column=4).fill = PatternFill("solid", fgColor=diffColor)
        colorizeRow(comparisonws, i+2, [1], coasterDict, coaster, manuColors)
    comparisonws.freeze_panes = comparisonws['A2']

    if useSpinner:
        spinner.stop()
    print(" ")



# ==================================================
#  OK, let's do this!
# ==================================================

if __name__ == "__main__": # allows us to put main at the beginning
    main()



# ==================================================
#  still to do
# ==================================================

# handle ties: decide which one wins, if possible
# make subsets: rankings of gigas, hypers, types, parks, etc

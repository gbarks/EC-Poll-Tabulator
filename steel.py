#!/usr/bin/env python3

# ==========================================================
#  ElloCoaster poll tabulator: steel coaster designers
#  Author: Grant Barker
# ==========================================================

from openpyxl.styles import PatternFill

designers = { # fill colors for prominent steel roller coaster manufacturers/designers
    "Bolliger & Mabillard"                  : PatternFill("solid", fgColor="fdb2b3"), # red
    "Maurer Rides GmbH"                     : PatternFill("solid", fgColor="fed185"), # orange
    "Arrow Dynamics"                        : PatternFill("solid", fgColor="fffd87"), # yellow
    "Mack Rides GmbH & Co KG"               : PatternFill("solid", fgColor="cde4cd"), # green
    "Intamin Amusement Rides"               : PatternFill("solid", fgColor="b2b4fd"), # blue
    "Gerstlauer Amusement Rides GmbH"       : PatternFill("solid", fgColor="c9b3d8"), # purple
    "Schwarzkopf"                           : PatternFill("solid", fgColor="fecdfe"), # pink
    "Rocky Mountain Construction"           : PatternFill("solid", fgColor="ceffff"), # cyan
    "Premier Rides"                         : PatternFill("solid", fgColor="e8d9c6"), # light brown
    "S&S Sansei Technologies"               : PatternFill("solid", fgColor="cb999a"), # dark brown
    "Vekoma"                                : PatternFill("solid", fgColor="8fc6d3"), # amber
    "Other Known Manufacturer"              : PatternFill("solid", fgColor="cccccc"), # dark gray
    ""                                      : PatternFill("solid", fgColor="eeeeee") # light gray
}

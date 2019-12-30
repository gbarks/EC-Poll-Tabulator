#!/usr/bin/env python3

# ==========================================================
#  ElloCoaster detailed_ballot reader
#  Author: Grant Barker
#
#  Requires Python 3
# ==========================================================

from coaster import Coaster
from openpyxl import load_workbook

def blank_to_none(string):
    if string:
        return string
    else:
        return None

def read_detailed_ballot(filepath, id_col=7, name_col=2, park_col=4):
    coasters = {}
    wb = load_workbook(filepath)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2):#, max_row=6): 
        rcid = row[id_col].value
        url = None
        if "HYPERLINK" in row[id_col].value:
            rcid = row[id_col].value[:-2].split('"')[-1]
            url = row[id_col].value.split('"')[1]

        c = Coaster(rcid,                         #rcid
                    row[name_col].value,          #name
                    row[park_col].value,          #park
                    url,                          #url
                    blank_to_none(row[3].value),  #altname
                    blank_to_none(row[5].value),  #country
                    blank_to_none(row[6].value),  #fullcity
                    blank_to_none(row[8].value),  #location
                    blank_to_none(row[9].value),  #state
                    blank_to_none(row[10].value), #city
                    blank_to_none(row[11].value), #status
                    blank_to_none(row[12].value), #opendate
                    blank_to_none(row[13].value), #closedate
                    blank_to_none(row[14].value), #rctype
                    blank_to_none(row[15].value), #scale
                    blank_to_none(row[16].value), #make
                    blank_to_none(row[17].value), #model
                    blank_to_none(row[18].value), #submodel
                    blank_to_none(row[19].value)) #tracks
        coasters[rcid] = c
        #print(vars(c))

    return coasters
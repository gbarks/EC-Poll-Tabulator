#!/usr/bin/env python

from math import sqrt

def getcolor(f, i=128, multiplier=1.0, offset = 0.0):
    f *= multiplier
    f += offset
    if f >= 1:
        f = f % 1
    if i >= 256:
        i = i % 256

    space = 255 - i

    if f < 0.1666666666:
        scalar = f * 6
        r = 255
        g = int(round(i + scalar * space))
        b = i

    elif f < 0.333333333:
        scalar = (f - 0.1666666666) * 6
        r = int(round(255 - scalar * space))
        g = 255
        b = i

    elif f < 0.5:
        scalar = (f - 0.333333333) * 6
        r = i
        g = 255
        b = int(round(i + scalar * space))

    elif f < 0.666666666:
        scalar = (f - 0.5) * 6
        r = i
        g = int(round(255 - scalar * space))
        b = 255

    elif f < 0.8333333333:
        scalar = (f - 0.666666666) * 6
        r = int(round(i + scalar * space))
        g = i
        b = 255

    else:
        scalar = (f - 0.8333333333) * 6
        r = 255
        g = i
        b = int(round(255 - scalar * space))

    colorstr = hex(r)[2:] + hex(g)[2:] + hex(b)[2:]
    print str(f) + "\t" + colorstr
    return colorstr

def phicolor(n, lightness=128, multiplier=1.0, offset=0.0):
    phi = float((1 + sqrt(5)) / 2)
    return getcolor(float(n * phi), lightness, multiplier, offset)

import sys
from random import random, seed
from openpyxl import Workbook
from openpyxl.styles import PatternFill

xl = Workbook()
ws = xl.active
seed()
offset = 0.4
multi = 1
for i in range(1, 100):
    ws.append(["test" + str(i)])
    ws["A" + str(i)].fill = PatternFill("solid", fgColor=phicolor(i, 186, multi, offset))
xl.save("Color Test.xlsx")
